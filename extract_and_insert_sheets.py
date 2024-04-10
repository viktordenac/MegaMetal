import os
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter as tk
from tkinter import simpledialog

def extract_list_from_file(file_path, list_name):
    workbook = load_workbook(filename=file_path, read_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet.title.lower() == list_name.lower():
            return list(sheet.values)

def insert_list_into_file(list_data, output_file, sheet_name):
    if not os.path.exists(output_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for row in list_data:
            sheet.append(row)
        workbook.save(output_file)
    else:
        workbook = load_workbook(filename=output_file)
        sheet = workbook.create_sheet(title=sheet_name)
        for row in list_data:
            sheet.append(row)
        workbook.save(output_file)

def extract_and_insert_lists(directory, output_file, list_name):
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            list_data = extract_list_from_file(file_path, list_name)
            if list_data:
                insert_list_into_file(list_data, output_file, list_name)
                print(f"List '{list_name}' from {filename} inserted into {output_file} as '{list_name}_{filename}'")

# Define the directory containing your Excel files
directory = "C:/Users/mm.student/Documents/Nekaj"

# Define the output file where the combined lists will be inserted
output_file = "C:/Users/mm.student/Desktop/output_file.xlsx"

# Specify the name of the list you want to extract and insert
list_name = "REALIZACIJA"

# Call the function to extract and insert lists
extract_and_insert_lists(directory, output_file, list_name)
